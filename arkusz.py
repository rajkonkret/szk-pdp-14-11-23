import openpyxl

# pip install openpyxl

# bibliotek słuząca do pracy z plikami np.: excel

workbook = openpyxl.load_workbook('sales.xlsx')
worksheet = workbook.active
print(worksheet)  # <Worksheet "Arkusz1">

lista = []

for w in worksheet:
    print(w)  # (<Cell 'Arkusz1'.A5>, <Cell 'Arkusz1'.B5>, <Cell 'Arkusz1'.C5>)

for i in range(0, worksheet.max_row):
    for col in worksheet.iter_cols(1, worksheet.max_column):
        lista.append(col[i].value)

print(lista)
# ['Sales Date', 'Sales Person', 'Amount', datetime.datetime(2018, 5, 12, 0, 0), 'Sila Ahmed', 60000,
#  datetime.datetime(2019, 12, 6, 0, 0), 'Mir Hossain', 50000, datetime.datetime(2020, 8, 9, 0, 0), 'Sarmin Jahan', 45000,
#  datetime.datetime(2021, 4, 7, 0, 0), 'Mahmudul Hasan', 30000]
print(lista[0:3])  # ['Sales Date', 'Sales Person', 'Amount']

for i in range(0, len(lista), 3):  # start, stop. step - krok
    print(lista[i:i + 3])
# [datetime.datetime(2018, 5, 12, 0, 0), 'Sila Ahmed', 60000]
# [datetime.datetime(2019, 12, 6, 0, 0), 'Mir Hossain', 50000]
# [datetime.datetime(2020, 8, 9, 0, 0), 'Sarmin Jahan', 45000]
# [datetime.datetime(2021, 4, 7, 0, 0), 'Mahmudul Hasan', 30000]
